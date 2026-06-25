# This file generates a detailed Excel report comparing raw and clean data.
# It uses the openpyxl library to write multiple spreadsheets.
# It includes summaries, column statistics, applied steps, and data samples.
# It saves the spreadsheet to a memory buffer and returns it as raw bytes.
# We do this to let users download a full audit report of all changes.

import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pandas as pd

def create_comparison_excel(raw_df, clean_df, execution_log):
    # This function creates a full Excel workbook comparing the tables.
    # It adds sheets for Summary, Column Stats, Applied Operations, and Samples.
    # It returns the workbook as bytes so it can be downloaded.
    wb = openpyxl.Workbook()
    
    # Define styles for a professional look
    title_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # 1. Summary sheet creation
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    ws_summary.append([])
    ws_summary.append(["Data Cleaning Summary Report"])
    ws_summary.cell(2, 1).font = title_font
    ws_summary.append([])
    
    # Retrieve stats for before and after cleaning
    raw_rows = len(raw_df)
    raw_cols = len(raw_df.columns)
    raw_dups = raw_df.duplicated().sum()
    raw_missing = raw_df.isna().sum().sum()
    
    clean_rows = len(clean_df)
    clean_cols = len(clean_df.columns)
    clean_dups = clean_df.duplicated().sum()
    clean_missing = clean_df.isna().sum().sum()
    
    ws_summary.append(["Metric", "Raw Dataset", "Cleaned Dataset", "Difference"])
    for col_idx in range(1, 5):
        cell = ws_summary.cell(ws_summary.max_row, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    metrics_data = [
        ("Total Rows", raw_rows, clean_rows, clean_rows - raw_rows),
        ("Total Columns", raw_cols, clean_cols, clean_cols - raw_cols),
        ("Duplicate Rows", raw_dups, clean_dups, clean_dups - raw_dups),
        ("Total Missing Cells", raw_missing, clean_missing, clean_missing - raw_missing)
    ]
    
    for label, raw_val, clean_val, diff in metrics_data:
        ws_summary.append([label, raw_val, clean_val, diff])
        row_idx = ws_summary.max_row
        ws_summary.cell(row_idx, 1).font = bold_font
        for col_idx in range(1, 5):
            cell = ws_summary.cell(row_idx, col_idx)
            cell.border = thin_border
            if col_idx > 1:
                cell.alignment = Alignment(horizontal="right")
                
    # 2. Column Stats sheet creation
    ws_cols = wb.create_sheet(title="Column Stats")
    ws_cols.append([])
    ws_cols.append(["Column Statistics Comparison"])
    ws_cols.cell(2, 1).font = title_font
    ws_cols.append([])
    
    # Write details for original raw columns
    ws_cols.append(["Raw Dataset Columns"])
    ws_cols.cell(ws_cols.max_row, 1).font = bold_font
    ws_cols.append(["Column Name", "Data Type", "Missing Count", "Missing %", "Unique Count"])
    for col_idx in range(1, 6):
        cell = ws_cols.cell(ws_cols.max_row, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for col_name in raw_df.columns:
        series = raw_df[col_name]
        missing = series.isna().sum()
        missing_pct = (missing / raw_rows * 100) if raw_rows > 0 else 0.0
        unique = series.nunique()
        ws_cols.append([col_name, str(series.dtype), missing, f"{missing_pct:.1f}%", unique])
        row_idx = ws_cols.max_row
        for col_idx in range(1, 6):
            cell = ws_cols.cell(row_idx, col_idx)
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right")
                
    ws_cols.append([])
    # Write details for cleaned columns
    ws_cols.append(["Cleaned Dataset Columns"])
    ws_cols.cell(ws_cols.max_row, 1).font = bold_font
    ws_cols.append(["Column Name", "Data Type", "Missing Count", "Missing %", "Unique Count"])
    for col_idx in range(1, 6):
        cell = ws_cols.cell(ws_cols.max_row, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for col_name in clean_df.columns:
        series = clean_df[col_name]
        missing = series.isna().sum()
        missing_pct = (missing / clean_rows * 100) if clean_rows > 0 else 0.0
        unique = series.nunique()
        ws_cols.append([col_name, str(series.dtype), missing, f"{missing_pct:.1f}%", unique])
        row_idx = ws_cols.max_row
        for col_idx in range(1, 6):
            cell = ws_cols.cell(row_idx, col_idx)
            cell.border = thin_border
            if col_idx >= 3:
                cell.alignment = Alignment(horizontal="right")
                
    # 3. Operations Log sheet creation
    ws_ops = wb.create_sheet(title="Operations Log")
    ws_ops.append([])
    ws_ops.append(["Applied Operations Log"])
    ws_ops.cell(2, 1).font = title_font
    ws_ops.append([])
    
    ws_ops.append(["Step #", "Operation", "Column", "Parameters", "Status", "Reason / Message"])
    for col_idx in range(1, 7):
        cell = ws_ops.cell(ws_ops.max_row, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for idx, log_entry in enumerate(execution_log):
        step_num = idx + 1
        op = log_entry.get("op", "")
        col = log_entry.get("column", "") or "Whole Table"
        params = str(log_entry.get("params", {}))
        status = log_entry.get("status", "")
        msg = log_entry.get("message", "") or log_entry.get("reason", "")
        
        ws_ops.append([step_num, op, col, params, status, msg])
        row_idx = ws_ops.max_row
        
        status_cell = ws_ops.cell(row_idx, 5)
        if status == "applied":
            status_cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        elif status == "skipped":
            status_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        elif status == "error":
            status_cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
            
        for col_idx in range(1, 7):
            cell = ws_ops.cell(row_idx, col_idx)
            cell.border = thin_border
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
                
    # 4. Raw Sample sheet creation
    ws_raw_sample = wb.create_sheet(title="Raw Sample")
    ws_raw_sample.append(["Raw Dataset Sample (First 100 rows)"])
    ws_raw_sample.cell(1, 1).font = bold_font
    ws_raw_sample.append(list(raw_df.columns))
    for col_idx in range(1, len(raw_df.columns) + 1):
        cell = ws_raw_sample.cell(2, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        
    for _, row in raw_df.head(100).iterrows():
        row_vals = [str(val) if not pd.isna(val) else "" for val in row]
        ws_raw_sample.append(row_vals)
        for col_idx in range(1, len(row_vals) + 1):
            ws_raw_sample.cell(ws_raw_sample.max_row, col_idx).border = thin_border
            
    # 5. Clean Sample sheet creation
    ws_clean_sample = wb.create_sheet(title="Clean Sample")
    ws_clean_sample.append(["Cleaned Dataset Sample (First 100 rows)"])
    ws_clean_sample.cell(1, 1).font = bold_font
    ws_clean_sample.append(list(clean_df.columns))
    for col_idx in range(1, len(clean_df.columns) + 1):
        cell = ws_clean_sample.cell(2, col_idx)
        cell.font = header_font
        cell.fill = header_fill
        
    for _, row in clean_df.head(100).iterrows():
        row_vals = [str(val) if not pd.isna(val) else "" for val in row]
        ws_clean_sample.append(row_vals)
        for col_idx in range(1, len(row_vals) + 1):
            ws_clean_sample.cell(ws_clean_sample.max_row, col_idx).border = thin_border
            
    # Auto adjust column widths for all sheets based on cell contents
    for ws in wb.worksheets:
        for col in ws.columns:
            max_len = 0
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)
            
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file.getvalue()
